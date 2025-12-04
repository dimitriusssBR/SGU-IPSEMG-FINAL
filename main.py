import datetime
import os
import re
import time
import logging
from logging.handlers import RotatingFileHandler
from typing import List, Optional
import nest_asyncio
from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel, validator
import openpyxl
from openpyxl.cell.cell import MergedCell
import textwrap
import unicodedata
from fuzzywuzzy import fuzz
from converte_em_pdf import gerar_pdf_final
import uuid
from pathlib import Path
from fastapi.responses import FileResponse
import tempfile
from fastapi.middleware.cors import CORSMiddleware

nest_asyncio.apply()
app = FastAPI()

# üîì CORS totalmente liberado (somente para desenvolvimento, mudar depois quando tiver dominio)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # Libera qualquer origem
    allow_credentials=True,
    allow_methods=["*"],          # Libera todos os m√©todos (GET, POST, DELETE, etc.)
    allow_headers=["*"],          # Libera todos os headers
)
VERSAO = "1.0 - 04/12/25"
MODO_DEBUG = None

IPSEMG_SADT = "IPSEMG_SADT.xlsx"
IPSEMG_INTERNACAO = "IPSEMG_INTERNACAO.xlsx"


# Carregar arquivo IPSEMG TXT
dados_ipsemg_normalizados = []

# -----------------------------------------------
# LOGGING
# -----------------------------------------------

logger = logging.getLogger("IPSEMG")
logger.setLevel(logging.INFO)

formatter = logging.Formatter(
    "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)

console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)

file_handler = RotatingFileHandler(
    "IPSEMG.log",
    maxBytes=5 * 1024 * 1024,
    backupCount=3,
    encoding="utf-8",
)
file_handler.setFormatter(formatter)

if not logger.handlers:
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

# -----------------------------------------------
# MODELOS
# -----------------------------------------------

class VersaoResponse(BaseModel):
    versao: str

class IpsemgPayload(BaseModel):
    nome_beneficiario: str
    operadora: Optional[str] = None
    prestador: Optional[str] = None
    matricula: Optional[str] = None
    sexo: Optional[str] = None
    uf: Optional[str] = None
    especialidade: Optional[str] = None
    crm: Optional[str] = None
    carater: str
    cid: Optional[str] = None
    solicitante: str
    indicacao_clinica: str
    tratamentos_realizados: Optional[str] = None
    hipotese: Optional[str] = None
    codigos: List[str]
    descricao: List[str]
    quantidades: List[int]
    tipo_internacao: Optional[str] = None
    regime: Optional[str] = None
    codigo_operadora: Optional[str] = None
    data_nascimento: Optional[str] = None
    assinatura: str
    data: str


def normalizar_texto(texto: str) -> str:
    texto = texto.lower()

    # Substitui√ß√µes m√©dicas antes de remover acentos
    substituicoes_medicas = {
        "resson√¢ncia magn√©tica": "rm",
        "ressonancia magnetica": "rm",
        "ressonancia": "rm",
        "resson√¢ncia": "rm",
        "tomografia computadorizada": "tc",
        "tomografia": "tc",
        "ultrassonografia": "us",
        "ultrassom": "us",
        "raio-x": "rx",
        "eletrocardiograma": "ecg",
    }
    for chave, valor in substituicoes_medicas.items():
        texto = texto.replace(chave, valor)

    # Agora remove acentos
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')

    # Substituir caracteres especiais por espa√ßo
    substituicoes = {
        '‚Äì': ' ', '-': ' ', '‚Äî': ' ', '‚àí': ' ',
        ',': ' ', '.': ' ', ';': ' ', ':': ' ',
        '(': ' ', ')': ' ', '[': ' ', ']': ' ',
        '{': ' ', '}': ' ', '<': ' ', '>': ' ',
        '!': ' ', '?': ' ', '"': ' ', "'": ' ',
        '&': ' ', '@': ' ', '#': ' ', '$': ' ',
        '%': ' ', '*': ' ', '+': ' ', '=': ' ',
        '/': ' ', '\\': ' ', '|': ' '
    }
    for original, sub in substituicoes.items():
        texto = texto.replace(original, sub)

    # Remover espa√ßos extras
    texto = ' '.join(texto.split())

    return texto

def carregar_dados_cbhpm_ipsemg():
    global dados_ipsemg_normalizados
    if dados_ipsemg_normalizados:
        return  # j√° carregado

    try:
        with open("ipsemg_refatorado.txt", encoding="utf-8") as f:
            linhas = f.readlines()
        for linha in linhas:
            linha = linha.strip()
            match = re.match(r'^(\d{1,2}\.\d{2}\.\d{2}\.\d{2}-\d)\s+(.*)', linha)
            if match:
                codigo = match.group(1)
                descricao = match.group(2)
                descricao_normalizada = normalizar_texto(descricao)
                dados_ipsemg_normalizados.append({
                    'normalizado': descricao_normalizada,
                    'original': descricao,
                    'codigo': codigo
                })
        logger.info(f"Arquivo CODIGOS IPSEMG TXT carregado com {len(dados_ipsemg_normalizados)} entradas")
    except Exception as e:
        logger.error(f"Erro ao carregar arquivo CODIGOS IPSEMG TXT: {str(e)}")

# === Buscar CBHPM ===
def buscar_chbpm(exame: str):
    carregar_dados_cbhpm_ipsemg()
    try:
        tempo_inicio = time.time()
        logger.info(f"=== IN√çCIO DA BUSCA CODIGOS IPSEMG ===")
        logger.info(f"Termo original: {exame}")

        # Normalizar o texto de entrada
        tempo_normalizacao = time.time()
        exame = normalizar_texto(exame)
        logger.info(f"Termo normalizado: {exame}")

        # Busca por c√≥digo CHBPM normalizado
        tempo_busca_codigo = time.time()
        exame_strip = exame.strip()
        if re.fullmatch(r'\d{8}', exame_strip):
            for dado in dados_ipsemg_normalizados:
                codigo_normalizado = dado['codigo'].replace(".", "").replace("-", "")
                if codigo_normalizado == exame_strip:
                    logger.info(f"Tempo total: {time.time() - tempo_inicio:.4f}s")
                    return {
                        "consulta": exame,
                        "sugestoes": [{
                            "descricao": dado['original'],
                            "codigo": dado['codigo'],
                            "score": 100
                        }]
                    }
        logger.info(f"Tempo busca por c√≥digo: {time.time() - tempo_busca_codigo:.4f}s")

        # Busca por express√£o normalizada
        tempo_preparacao = time.time()
        stopwords = {"de", "do", "da", "e", "a", "o", "para", "por"}
        termo_normalizado = ' '.join([
            palavra for palavra in exame.split()
            if palavra not in stopwords
        ])

        # Busca fuzzy otimizada
        tempo_busca_fuzzy = time.time()
        resultados = []
        total_comparacoes = 0
        if termo_normalizado.startswith("diaria"):
            base_busca = [
                dado for dado in dados_ipsemg_normalizados
                if dado['normalizado'].startswith("diaria")
            ]
        else:
            base_busca = [
                dado for dado in dados_ipsemg_normalizados
                if termo_normalizado in dado['normalizado']
            ]

            if not base_busca:
                palavras = termo_normalizado.split()
                base_busca = [
                    dado for dado in dados_ipsemg_normalizados
                    if all(p in dado['normalizado'] for p in palavras)
                ]
        for dado in base_busca:
            total_comparacoes += 1
            score = int(fuzz.WRatio(termo_normalizado, dado['normalizado']))
            if score >= 70 or termo_normalizado in dado['normalizado']:
                resultados.append({
                    'descricao': dado['original'],
                    'codigo': dado['codigo'],
                    'score': score
                })

        # Ordenar resultados
        tempo_ordenacao = time.time()
        palavras_busca = set(termo_normalizado.split())

        def contem_todas_as_palavras(descricao):
            return palavras_busca.issubset(set(normalizar_texto(descricao).split()))

        resultados.sort(key=lambda x: (
            not contem_todas_as_palavras(x['descricao']),  # True vira 1, False vira 0 (queremos False primeiro)
            not x['descricao'].lower().startswith(termo_normalizado),
            -x['score']
        ))

        # Limitar a 20 resultados
        resultados = resultados[:5]
        logger.info(f"Total de resultados encontrados: {len(resultados)}")
        logger.info(f"Tempo total: {time.time() - tempo_inicio:.4f}s")
        logger.info("=== FIM DA BUSCA CBHPM ===")

        return {
            "consulta": exame,
            "sugestoes": resultados
        }

    except Exception as e:
        logger.info(f"Erro na busca CBHPM: {str(e)}", exc_info=True)
        return {"consulta": exame, "sugestoes": [], "erro": str(e)}


# === FastAPI Schemas ===
class CBHPMRequest(BaseModel):
    exame: str


# === Endpoint: Buscar CBHPM ===
@app.post("/buscar-chbpm")
async def buscar_chbpm_endpoint(request: CBHPMRequest):
    try:
        resultado = buscar_chbpm(request.exame)
        if isinstance(resultado, JSONResponse):
            return resultado
        return resultado
    except Exception as e:
        logger.info(f"Erro ao buscar CODIGO IPSEMG: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"mensagem": f"Erro ao buscar CODIGO IPSEMG: {str(e)}"}
        )

def remover_acentos(texto):
    return unicodedata.normalize('NFD', texto).encode('ascii', 'ignore').decode('utf-8')

def sanitize_filename_part(text: str) -> str:
    """
    Remove acentos, caracteres estranhos e troca espa√ßos por _.
    Deixa s√≥ letras, n√∫meros, _ e -.
    """
    if not text:
        return "Paciente"

    # remove acentos
    text_norm = unicodedata.normalize("NFKD", text)
    text_ascii = text_norm.encode("ascii", "ignore").decode("ascii")

    # troca tudo que n√£o √© letra/n√∫mero/_/- por _
    text_clean = re.sub(r"[^A-Za-z0-9_-]+", "_", text_ascii)

    # remove _ duplicado e bordas
    text_clean = re.sub(r"_+", "_", text_clean).strip("_")

    return text_clean or "Paciente"

def set_cell_value_safely(ws, coord: str, value):
    """
    Se a c√©lula for parte de um merged range, grava na c√©lula
    superior esquerda do merge. Se n√£o for, grava direto.
    """
    cell = ws[coord]

    if not isinstance(cell, MergedCell):
        ws[coord] = value
        return

    # Procura o range mesclado ao qual essa c√©lula pertence
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            top_left_coord = str(merged_range.coord).split(":")[0]
            ws[top_left_coord] = value
            return

    raise ValueError(f"C√©lula {coord} √© mesclada mas o range n√£o foi encontrado.")


async def _ipsemg_sadt_core(payload: IpsemgPayload) -> dict:
    if not os.path.exists(IPSEMG_SADT):
        raise HTTPException(status_code=500, detail=f"Arquivo {IPSEMG_SADT} n√£o encontrado")

    # ----- diret√≥rio isolado por requisi√ß√£o (se j√° estiver assim na sua rota atual, mantenha) -----
    request_id = uuid.uuid4().hex
    base_dir = Path("/tmp") / request_id
    base_dir.mkdir(parents=True, exist_ok=True)

    # caminhos exclusivos dessa requisi√ß√£o
    xlsx_path = base_dir / "ipsemg_sadt_output.xlsx"

    # carrega template
    wb = openpyxl.load_workbook(IPSEMG_SADT)
    ws = wb.active  # primeira aba

    # -------------------------------------------------
    # CAMPOS SIMPLES
    # -------------------------------------------------

    # nome_beneficiario -> B7
    set_cell_value_safely(ws, "B7", payload.nome_beneficiario)

    # solicitante -> B13
    set_cell_value_safely(ws, "B13", payload.solicitante)

    # prestador -> B10
    set_cell_value_safely(ws, "B10", payload.prestador)

    # matricula -> W10
    set_cell_value_safely(ws, "W10", payload.matricula)

    # uf -> B16
    set_cell_value_safely(ws, "B16", payload.uf)

    # crm -> Z13
    set_cell_value_safely(ws, "Z13", payload.crm)

    # especialidade -> G16
    set_cell_value_safely(ws, "G16", payload.especialidade)

    # cid -> Z20
    set_cell_value_safely(ws, "Z20", payload.cid)

    # assinatura -> J63
    set_cell_value_safely(ws, "J63", payload.assinatura)

    # data "dd/mm/aaaa" -> C63 (dia), E63 (mes), G63 (ano)
    dia, mes, ano = "", "", ""
    try:
        partes = (payload.data or "").strip().split("/")
        if len(partes) == 3:
            dia, mes, ano = partes
    except Exception:
        pass

    set_cell_value_safely(ws, "C63", dia)
    set_cell_value_safely(ws, "E63", mes)
    set_cell_value_safely(ws, "G63", ano)

    # -------------------------------------------------
    # CAR√ÅTER (ELETIVO / URG√äNCIA)
    # C20 se ELETIVO, I20 se URG√äNCIA
    # -------------------------------------------------
    car = (payload.carater or "").strip().lower()

    # Limpamos os dois primeiro
    set_cell_value_safely(ws, "C20", "")
    set_cell_value_safely(ws, "I20", "")

    if "elet" in car:
        set_cell_value_safely(ws, "C20", "X")
    elif "urg" in car or "√∫rg" in car:
        set_cell_value_safely(ws, "I20", "X")

    # -------------------------------------------------
    # INDICA√á√ÉO CL√çNICA -> C23 a C27 (com quebra em linhas)
    # -------------------------------------------------
    linhas_indicacao = textwrap.wrap(payload.indicacao_clinica or "", width=70)
    for i in range(5):  # C23..C27
        texto_linha = linhas_indicacao[i] if i < len(linhas_indicacao) else ""
        row = 23 + i
        coord = f"C{row}"
        set_cell_value_safely(ws, coord, texto_linha)

    # -------------------------------------------------
    # TRATAMENTOS REALIZADOS -> C30 a C34 (mesma regra)
    # -------------------------------------------------
    linhas_trat = textwrap.wrap(payload.tratamentos_realizados or "", width=70)
    for i in range(5):  # C30..C34
        texto_linha = linhas_trat[i] if i < len(linhas_trat) else ""
        row = 30 + i
        coord = f"C{row}"
        set_cell_value_safely(ws, coord, texto_linha)

    # -------------------------------------------------
    # HIP√ìTESE DIAGN√ìSTICA -> C37 a C41 (mesma regra)
    # -------------------------------------------------
    linhas_hipotese = textwrap.wrap(payload.hipotese or "", width=70)
    for i in range(5):  # C37..C41
        texto_linha = linhas_hipotese[i] if i < len(linhas_hipotese) else ""
        row = 37 + i
        coord = f"C{row}"
        set_cell_value_safely(ws, coord, texto_linha)

    # -------------------------------------------------
    # C√ìDIGOS / DESCRI√á√ÉO / QUANTIDADES
    # C√ìDIGOS -> B46 a B52
    # DESCRI√á√ÉO -> G46 a G52
    # QUANTIDADES -> AE46 a AE52
    # -------------------------------------------------
    max_linhas = 7  # 46..52 = 7 linhas

    codigos = payload.codigos or []
    quantidades = payload.quantidades or []
    descricoes = getattr(payload, "descricao", None) or []

    for idx in range(max_linhas):
        row = 46 + idx

        cod_coord = f"B{row}"
        desc_coord = f"G{row}"
        qtd_coord = f"AE{row}"

        set_cell_value_safely(ws, cod_coord, codigos[idx] if idx < len(codigos) else "")
        set_cell_value_safely(ws, desc_coord, descricoes[idx] if idx < len(descricoes) else "")
        set_cell_value_safely(ws, qtd_coord, quantidades[idx] if idx < len(quantidades) else "")

    # salva o XLSX desta requisi√ß√£o

    wb.save(xlsx_path)

    try:
        pdf_file = gerar_pdf_final(str(xlsx_path))
    except Exception as e:
        pdf_file = None
        print(f"Erro ao converter para PDF: {e}")

    return {
        "status": "ok",
        "mensagem": "GUIA IPSEMG SADT preenchida com sucesso",
        "arquivo_xlsx": str(xlsx_path),
        "arquivo_pdf": pdf_file,
        "payload": payload.model_dump()
    }

@app.post("/ipsemg-sadt")
async def ipsemg_sadt(payload: IpsemgPayload):
    return await _ipsemg_sadt_core(payload)

async def _ipsemg_internacao_core(payload: IpsemgPayload) -> dict:
    if not os.path.exists(IPSEMG_INTERNACAO):
        raise HTTPException(
            status_code=500,
            detail=f"Arquivo {IPSEMG_INTERNACAO} n√£o encontrado"
        )

    # ----- diret√≥rio isolado por requisi√ß√£o -----
    request_id = uuid.uuid4().hex
    base_dir = Path("/tmp") / request_id
    base_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = base_dir / "ipsemg_internacao_output.xlsx"

    wb = openpyxl.load_workbook(IPSEMG_INTERNACAO)
    ws = wb.active  # primeira aba

    # -------------------------------------------------
    # CAR√ÅTER (ELETIVO / URG√äNCIA)
    # -------------------------------------------------
    car = (payload.carater or "").strip().lower()

    set_cell_value_safely(ws, "E9", "")
    set_cell_value_safely(ws, "R9", "")

    if "elet" in car:
        set_cell_value_safely(ws, "E9", "X")
    elif "urg" in car or "√∫rg" in car:
        set_cell_value_safely(ws, "R9", "X")

    # -------------------------------------------------
    # CAMPOS SIMPLES
    # -------------------------------------------------
    set_cell_value_safely(ws, "B13", payload.matricula)      # matr√≠cula
    set_cell_value_safely(ws, "H13", payload.prestador)      # prestador

    sexo = (payload.sexo or "").strip().lower()

    set_cell_value_safely(ws, "U17", "")
    set_cell_value_safely(ws, "AB17", "")

    if "masc" in sexo:
        set_cell_value_safely(ws, "U17", "X")
    elif "fem" in sexo:
        set_cell_value_safely(ws, "AB17", "X")

    # nome benefici√°rio -> B20
    set_cell_value_safely(ws, "B20", payload.nome_beneficiario)

    # -------------------------------------------------
    # DATA NASCIMENTO BENEFICI√ÅRIO (dd/mm/aaaa)
    # -------------------------------------------------
    dia_nasc, mes_nasc, ano_nasc = "", "", ""
    try:
        if payload.data_nascimento:
            partes = payload.data_nascimento.strip().split("/")
            if len(partes) == 3:
                dia_nasc, mes_nasc, ano_nasc = partes
    except Exception:
        pass

    set_cell_value_safely(ws, "L17", dia_nasc)
    set_cell_value_safely(ws, "N17", mes_nasc)
    set_cell_value_safely(ws, "P17", ano_nasc)

    # -------------------------------------------------
    # C√ìDIGOS / DESCRI√á√ïES / QUANTIDADES
    # -------------------------------------------------
    max_linhas = 10  # 29..38

    codigos = payload.codigos or []
    descricoes = getattr(payload, "descricoes", None) or getattr(payload, "descricao", None) or []
    quantidades = payload.quantidades or []

    for idx in range(max_linhas):
        row = 29 + idx

        cod_coord = f"B{row}"
        desc_coord = f"G{row}"
        qtd_coord  = f"AG{row}"

        set_cell_value_safely(ws, cod_coord, codigos[idx]      if idx < len(codigos)      else "")
        set_cell_value_safely(ws, desc_coord, descricoes[idx]  if idx < len(descricoes)   else "")
        set_cell_value_safely(ws, qtd_coord,  quantidades[idx] if idx < len(quantidades)  else "")

    # -------------------------------------------------
    # INDICA√á√ÉO CL√çNICA -> B53
    # -------------------------------------------------
    set_cell_value_safely(ws, "B53", payload.indicacao_clinica)

    # -------------------------------------------------
    # HIP√ìTESE + CID -> B59
    # -------------------------------------------------
    hipotese = (payload.hipotese or "").strip()
    cid = (payload.cid or "").strip()

    if hipotese and cid:
        texto_hipotese = f"{hipotese} - {cid}"
    elif hipotese:
        texto_hipotese = hipotese
    else:
        texto_hipotese = cid

    set_cell_value_safely(ws, "B59", texto_hipotese)

    # -------------------------------------------------
    # SOLICITANTE / CRM / ESPECIALIDADE
    # -------------------------------------------------
    set_cell_value_safely(ws, "B62", payload.solicitante)
    set_cell_value_safely(ws, "B64", payload.crm)
    set_cell_value_safely(ws, "I64", payload.especialidade)

    # -------------------------------------------------
    # DATA DA GUIA (dd/mm/aaaa) -> AB64, AD64, AF64
    # -------------------------------------------------
    dia, mes, ano = "", "", ""
    try:
        if payload.data:
            partes = payload.data.strip().split("/")
            if len(partes) == 3:
                dia, mes, ano = partes
    except Exception:
        pass

    set_cell_value_safely(ws, "AB64", dia)
    set_cell_value_safely(ws, "AD64", mes)
    set_cell_value_safely(ws, "AF64", ano)

    # assinatura -> C67
    set_cell_value_safely(ws, "C67", payload.assinatura)

    # salva o XLSX desta requisi√ß√£o
    wb.save(xlsx_path)

    try:
        pdf_file = gerar_pdf_final(str(xlsx_path))
    except Exception as e:
        pdf_file = None
        print(f"Erro ao converter para PDF: {e}")

    return {
        "status": "ok",
        "mensagem": "GUIA IPSEMG INTERNACAO preenchida com sucesso",
        "arquivo_xlsx": str(xlsx_path),
        "arquivo_pdf": pdf_file,
        "payload": payload.model_dump()
    }


# Endpoint JSON (mant√©m compatibilidade com o que j√° existe)
@app.post("/ipsemg-internacao")
async def ipsemg_internacao(payload: IpsemgPayload):
    return await _ipsemg_internacao_core(payload)

@app.post("/ipsemg-sadt-saas")
async def ipsemg_sadt_saas(payload: IpsemgPayload):
    # Reusa a MESMA l√≥gica que j√° sabemos que funciona
    result = await _ipsemg_sadt_core(payload)

    pdf_path = result.get("arquivo_pdf")
    if not pdf_path or not Path(pdf_path).exists():
        raise HTTPException(status_code=500, detail="PDF n√£o encontrado ap√≥s gera√ß√£o da guia")

    # Monta o nome do arquivo: LIA_Sgu_Express_+nome_beneficiario+IPSEMG+data.pdf
    nome_benef = sanitize_filename_part(payload.nome_beneficiario or "Paciente")

    if getattr(payload, "data", None):
        data_bruta = payload.data
    else:
        data_bruta = datetime.now().strftime("%d/%m/%Y")

    data_formatada = data_bruta.replace("/", "-")

    filename = f"LIA_Sgu_Express_+{nome_benef}+IPSEMG+{data_formatada}.pdf"

    return FileResponse(
        path=pdf_path,
        media_type="application/pdf",
        filename=filename
    )

@app.post("/ipsemg-internacao-saas")
async def ipsemg_internacao_saas(payload: IpsemgPayload):
    # Reusa a MESMA l√≥gica que j√° sabemos que funciona
    result = await _ipsemg_internacao_core(payload)

    pdf_path = result.get("arquivo_pdf")
    if not pdf_path or not Path(pdf_path).exists():
        raise HTTPException(
            status_code=500,
            detail="PDF n√£o encontrado ap√≥s gera√ß√£o da guia de interna√ß√£o"
        )

    # Monta o nome do arquivo: LIA_Sgu_Express_+nome_beneficiario+IPSEMG+data.pdf
    nome_benef = sanitize_filename_part(payload.nome_beneficiario or "Paciente")

    if getattr(payload, "data", None):
        data_bruta = payload.data
    else:
        data_bruta = datetime.datetime.now().strftime("%d/%m/%Y")

    data_formatada = data_bruta.replace("/", "-")

    filename = f"LIA_Sgu_Express_+{nome_benef}+IPSEMG+{data_formatada}.pdf"

    return FileResponse(
        path=pdf_path,
        media_type="application/pdf",
        filename=filename
    )

@app.get("/versao", response_model=VersaoResponse)
async def versao():
    logger.info("Endpoint /versao chamado")
    return {"versao": VERSAO}

# -----------------------------------------------
# RODAR LOCALMENTE / CLOUD RUN
# -----------------------------------------------

if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
